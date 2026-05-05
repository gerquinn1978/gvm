"""File loaders and encryption refusal for /gvm-analysis (AN-2, AN-42, AN-43).

This module is the single input boundary for the analysis engine. It accepts
xlsx, xls, csv, tsv, parquet, json, and jsonl files; every other format is
rejected with a diagnostic naming the supported extensions (AN-2 / NFR-3).

Per cross-cutting ADR-007 this module follows SRP rigidly:

* It does NOT aggregate sheets or files — that lives in ``aggregation.py``.
* It does NOT detect AN-40 already-anonymised tokens — that lives in
  ``token_detect.py``.
* It does NOT compute dtype coercions or type drift — ``type_drift.py`` owns
  that.

``load`` always returns a pandas DataFrame. Polars may be used internally as
a lazy scan for > 100 MB files (performance tier), but the public return
type is pandas (ADR-001 fallback protocol / design-review-003 L-3).

Encrypted xlsx files are refused outright per AN-42: no password prompting,
no decryption library invocation. A real encrypted xlsx is a Compound File
Binary Format (CFBF / OLE) container, so the zipfile-based openpyxl loader
raises :class:`zipfile.BadZipFile`. When that happens we inspect the first
eight bytes for the OLE magic header (``D0 CF 11 E0 A1 B1 1A E1``) and
raise :class:`~_shared.diagnostics.EncryptedFileError`; genuine corruption
raises :class:`~_shared.diagnostics.MalformedFileError` instead. Both the
``load`` and ``extract_formulas`` paths route through the same
``_open_workbook`` helper so the encryption-refusal contract applies
uniformly across every xlsx entry point (including the provenance formula
extraction consumed by P2-C03).

Text-format parser failures (truncated CSV, malformed JSON) are caught at
this layer and re-raised as :class:`MalformedFileError` with
``kind="parser_error"`` so the diagnostic formatter in P2-C04 can produce
format-agnostic guidance (cross-cutting.md §890-892).
"""

from __future__ import annotations

import importlib.util
import json as _json
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd

from _shared.diagnostics import EncryptedFileError, MalformedFileError

__all__ = ["load", "extract_formulas", "check_dependencies"]


_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_SUPPORTED_EXTENSIONS: tuple[str, ...] = (
    "xlsx",
    "xls",
    "csv",
    "tsv",
    "parquet",
    "json",
    "jsonl",
)

# Importable package names, not pip-installable names. Orchestration
# (P5-C01) maps these back to pip names when composing the install command.
# Order is canonical and load-bearing: it keeps diagnostic output stable
# across invocations so users get the same message for the same missing
# environment.
_REQUIRED_IMPORT_NAMES: tuple[str, ...] = (
    "pandas",
    "numpy",
    "scipy",
    "sklearn",
    "statsmodels",
    "openpyxl",
)


def load(path: Path | str, *, sheet: str | None = None) -> pd.DataFrame:
    """Load a file into a pandas DataFrame, dispatching by extension.

    Raises
    ------
    FileNotFoundError
        When the path does not exist on disk.
    ValueError
        When the extension is unsupported; when a multi-sheet xlsx is loaded
        without ``sheet=`` (the caller is directed to ``aggregation.py``);
        when ``sheet=`` names a sheet that does not exist in the workbook.
    EncryptedFileError
        When an xlsx file is encrypted (CFBF/OLE container).
    MalformedFileError
        When a text-format file is truncated or unparsable
        (``kind="parser_error"``); when the file is readable as bytes but
        has no usable data rows (``kind="no_data_rows"``); when an xlsx is
        structurally corrupt (``kind="corrupt_xlsx"``); when a legacy
        ``.xls`` file is supplied but the optional ``xlrd`` engine is not
        installed (``kind="xls_engine_missing"``).
    OSError
        Permission denied and other I/O failures propagate unchanged.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    suffix = path.suffix.lower().lstrip(".")
    if suffix not in _SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"unsupported file extension for {path}: .{suffix} "
            f"(supported: {', '.join(_SUPPORTED_EXTENSIONS)})"
        )

    if suffix == "xlsx":
        return _load_xlsx(path, sheet=sheet)
    if suffix == "xls":
        return _load_xls(path, sheet=sheet)
    if suffix == "csv":
        return _load_delimited(path, sep=",")
    if suffix == "tsv":
        return _load_delimited(path, sep="\t")
    if suffix == "parquet":
        return _load_parquet(path)
    if suffix == "json":
        return _load_json(path, lines=False)
    if suffix == "jsonl":
        return _load_json(path, lines=True)

    # Unreachable: _SUPPORTED_EXTENSIONS is closed over the dispatch table
    # above. Left as a defensive raise so a future edit that extends the
    # extension tuple without extending the dispatcher fails loudly.
    raise ValueError(f"no loader registered for extension: .{suffix}")


def extract_formulas(
    path: Path | str, *, sheet: str | None = None
) -> list[dict[str, str]]:
    """Return ``[{column, formula}]`` entries for xlsx formula columns.

    AN-43: the engine reads cached numeric values for analysis (see
    :func:`load`) and records the formula strings verbatim in the
    provenance footer. This function is pass 2 of the two-pass openpyxl
    pattern: ``data_only=False``, which returns the formula text rather
    than the cached value.

    For non-xlsx extensions returns ``[]`` — no other format encodes
    formulas in a way the engine honours.

    Routes through :func:`_open_workbook` so an encrypted xlsx handed to
    this path raises :class:`EncryptedFileError` identically to
    :func:`load` (AN-42 refusal contract).
    """
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        return []

    wb = _open_workbook(path, data_only=False)
    ws = wb[sheet] if sheet is not None else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=False)

    header_row = next(rows, None)
    if header_row is None:
        return []
    header = [c.value for c in header_row]

    first_data_row = next(rows, None)
    if first_data_row is None:
        return []

    formulas: list[dict[str, str]] = []
    for col_idx, cell in enumerate(first_data_row):
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            column_name = header[col_idx] if col_idx < len(header) else None
            if column_name is not None:
                formulas.append({"column": str(column_name), "formula": value})
    return formulas


def check_dependencies() -> list[str]:
    """Return the list of required-but-missing top-level packages.

    Uses :func:`importlib.util.find_spec` via the :func:`_spec_finder`
    wrapper, which inspects the import machinery without triggering a real
    import. Returns package names by their IMPORTABLE identifier (e.g.
    ``sklearn`` not ``scikit-learn``) — the orchestration layer maps back
    to pip names when composing an install command.

    Tests monkeypatch ``_spec_finder`` rather than
    ``importlib.util.find_spec`` so their fallthrough branches can call
    the real finder without recursing into their own patch.
    """
    missing: list[str] = []
    for name in _REQUIRED_IMPORT_NAMES:
        if _spec_finder(name) is None:
            missing.append(name)
    return missing


def _spec_finder(name: str) -> object | None:
    """Indirection around :func:`importlib.util.find_spec` for testing.

    Keeping this as a module-level function (rather than a direct reference
    to ``importlib.util.find_spec``) lets tests replace the resolution
    strategy for a single call site without monkeypatching the stdlib
    globally and triggering recursion in fallthrough branches.
    """
    return importlib.util.find_spec(name)


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------


def _open_workbook(path: Path, *, data_only: bool) -> Any:
    """Open an xlsx workbook with uniform encryption / corruption mapping.

    This is the single point at which openpyxl exceptions cross the module
    boundary — both :func:`load` (pass 1, cached values) and
    :func:`extract_formulas` (pass 2, formula strings) route through here,
    so the AN-42 encryption-refusal contract applies identically on every
    xlsx entry point.

    Uses :func:`isinstance` against the optionally-imported
    ``InvalidFileException`` rather than matching the class name as a
    string — that makes the mapping robust to openpyxl renaming its
    internal exception classes in a future release.
    """
    import openpyxl

    try:
        invalid_file_exc: type[Exception] | None = (
            openpyxl.utils.exceptions.InvalidFileException
        )
    except AttributeError:  # pragma: no cover — future openpyxl API shift
        invalid_file_exc = None

    try:
        return openpyxl.load_workbook(path, data_only=data_only, read_only=True)
    except zipfile.BadZipFile as exc:
        # Real encrypted xlsx files are CFBF/OLE containers, not zips.
        # Inspect the first 8 bytes to distinguish encryption from generic
        # zip corruption. A file that is not a zip AND not an OLE container
        # is treated as corrupt rather than encrypted — still refused, just
        # with a different exception type so the diagnostic formatter
        # (P2-C04) can produce the right guidance.
        with path.open("rb") as fh:
            head = fh.read(8)
        if head == _OLE_MAGIC:
            raise EncryptedFileError(path) from exc
        raise MalformedFileError(path, row=None, col=None, kind="corrupt_xlsx") from exc
    except Exception as exc:  # noqa: BLE001
        if invalid_file_exc is not None and isinstance(exc, invalid_file_exc):
            raise MalformedFileError(
                path, row=None, col=None, kind="corrupt_xlsx"
            ) from exc
        raise


def _load_xlsx(path: Path, *, sheet: str | None) -> pd.DataFrame:
    """Pass-1 xlsx loader (``data_only=True``, cached numeric values)."""
    wb = _open_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    if sheet is None:
        if len(sheet_names) > 1:
            raise ValueError(
                f"multi-sheet xlsx detected ({len(sheet_names)} sheets: "
                f"{', '.join(sheet_names)}); use _shared.aggregation to "
                "pick a sheet or aggregate across sheets"
            )
        sheet = sheet_names[0]
    elif sheet not in sheet_names:
        raise ValueError(
            f"sheet {sheet!r} not found in {path}; available sheets: "
            f"{', '.join(sheet_names)}"
        )

    try:
        frame = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    except (
        zipfile.BadZipFile
    ) as exc:  # pragma: no cover — pre-filtered by _open_workbook
        raise MalformedFileError(path, row=None, col=None, kind="corrupt_xlsx") from exc

    # Empty-sheet guard per AN-2 / TC-AN-2-04. ``pd.read_excel`` returns a
    # zero-row frame for a header-only sheet; surface a structured error so
    # the diagnostic formatter produces actionable guidance rather than a
    # downstream "no data" stacktrace.
    if len(frame) == 0:
        raise MalformedFileError(path, row=None, col=None, kind="no_data_rows")

    return frame


def _load_xls(path: Path, *, sheet: str | None) -> pd.DataFrame:
    """Legacy ``.xls`` loader via the optional ``xlrd`` engine.

    xlrd≥2.0 supports only ``.xls`` (legacy BIFF) — which is why we pin the
    engine here rather than letting pandas guess. ``.xls`` is an AN-2 MUST
    requirement so the engine absence is mapped to a structured
    :class:`MalformedFileError` with ``kind="xls_engine_missing"`` so the
    diagnostic formatter (P2-C04) can guide the user to ``pip install xlrd``
    rather than surfacing pandas' generic "engine not installed" error.
    """
    if _spec_finder("xlrd") is None:
        raise MalformedFileError(path, row=None, col=None, kind="xls_engine_missing")
    try:
        if sheet is None:
            frame = pd.read_excel(path, engine="xlrd")
        else:
            frame = pd.read_excel(path, sheet_name=sheet, engine="xlrd")
    except Exception as exc:  # noqa: BLE001
        raise MalformedFileError(path, row=None, col=None, kind="corrupt_xls") from exc

    if isinstance(frame, dict):
        # Multi-sheet .xls without sheet= — consistent with xlsx path.
        raise ValueError(
            f"multi-sheet xls detected ({len(frame)} sheets: "
            f"{', '.join(frame.keys())}); use _shared.aggregation to "
            "pick a sheet or aggregate across sheets"
        )

    if len(frame) == 0:
        raise MalformedFileError(path, row=None, col=None, kind="no_data_rows")
    return frame


def _load_delimited(path: Path, *, sep: str) -> pd.DataFrame:
    try:
        return pd.read_csv(path, sep=sep)
    except pd.errors.ParserError as exc:
        # pandas stashes the row in the exception message as
        # "Error tokenizing data. C error: Expected N fields in line M …".
        # Surface the raw message via the kind-specific MalformedFileError —
        # the formatter in P2-C04 extracts row/col if present.
        raise MalformedFileError(path, row=None, col=None, kind="parser_error") from exc
    except pd.errors.EmptyDataError as exc:
        raise MalformedFileError(path, row=None, col=None, kind="no_data_rows") from exc


def _load_parquet(path: Path) -> pd.DataFrame:
    try:
        return pd.read_parquet(path)
    except (OSError, ValueError) as exc:
        # pyarrow raises OSError for corrupt parquet; pd wraps some cases in
        # ValueError. Both are structural malformations at this boundary.
        raise MalformedFileError(path, row=None, col=None, kind="parser_error") from exc


def _load_json(path: Path, *, lines: bool) -> pd.DataFrame:
    try:
        return pd.read_json(path, lines=lines)
    except (ValueError, _json.JSONDecodeError) as exc:
        raise MalformedFileError(path, row=None, col=None, kind="parser_error") from exc
